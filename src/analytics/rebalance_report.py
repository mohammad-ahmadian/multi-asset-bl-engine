import logging
from pathlib import Path
from typing import Dict, Any, Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class RebalanceReportGenerator:
    """
    Builds institutional-grade Excel Rebalance Order Sheets and Risk Factsheets.
    """

    # Institutional Color Palette
    NAVY_HEADER = "1B365D"
    SUBHEADER_FILL = "E8EEF5"
    BUY_FILL = "E6F4EA"
    BUY_FONT = "137333"
    SELL_FILL = "FCE8E6"
    SELL_FONT = "C5221F"
    BORDER_COLOR = "D0D5DD"

    def __init__(
        self,
        portfolio_aum: float = 25_000_000.0, # Default: $25 Million institutional mandate
        base_currency: str = "USD",
        fund_name: str = "Global Multi-Asset Tactical Fund (UCITS)"
    ):
        self.aum = portfolio_aum
        self.currency = base_currency
        self.fund_name = fund_name

    def generate_trade_orders_table(
        self,
        current_prices: pd.Series,
        current_weights: pd.Series,
        target_weights: pd.Series,
        asset_names: Dict[str, str],
        asset_classes: Dict[str, str],
        tx_cost_bps: float = 10.0
    ) -> pd.DataFrame:
        """
        Calculates exact monetary amounts and integer share quantities to buy/sell.
        """
        tickers = list(target_weights.index)
        records = []
        cost_rate = tx_cost_bps / 10000.0

        for t in tickers:
            px = float(current_prices.get(t, 0.0))
            w_curr = float(current_weights.get(t, 0.0))
            w_targ = float(target_weights.get(t, 0.0))

            val_curr = w_curr * self.aum
            val_targ = w_targ * self.aum

            shares_curr = int(val_curr / px) if px > 0 else 0
            shares_targ = int(val_targ / px) if px > 0 else 0

            delta_val = val_targ - val_curr
            delta_shares = shares_targ - shares_curr
            delta_w = w_targ - w_curr

            if delta_shares > 0:
                action = "BUY"
            elif delta_shares < 0:
                action = "SELL"
            else:
                action = "HOLD"

            est_cost = abs(delta_val) * cost_rate

            records.append({
                "ticker": t,
                "asset_name": asset_names.get(t, t),
                "asset_class": asset_classes.get(t, "Other"),
                "price": px,
                "current_weight": w_curr,
                "current_value": val_curr,
                "current_shares": shares_curr,
                "target_weight": w_targ,
                "target_value": val_targ,
                "target_shares": shares_targ,
                "weight_delta": delta_w,
                "action": action,
                "trade_value": abs(delta_val),
                "trade_shares": abs(delta_shares),
                "tx_cost": est_cost
            })

        return pd.DataFrame(records)

    def export_to_excel(
        self,
        trades_df: pd.DataFrame,
        risk_metrics: Dict[str, Any],
        asset_class_summary: pd.DataFrame,
        output_filepath: Path
    ) -> Path:
        """
        Creates a beautifully styled, publication-ready Excel workbook.
        """
        output_filepath.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rebalance & Order Sheet"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="E0E0E0")
        font_section = Font(name="Calibri", size=12, bold=True, color=self.NAVY_HEADER)
        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=10, bold=True)
        font_regular = Font(name="Calibri", size=10)
        
        fill_navy = PatternFill(start_color=self.NAVY_HEADER, end_color=self.NAVY_HEADER, fill_type="solid")
        fill_sub = PatternFill(start_color=self.SUBHEADER_FILL, end_color=self.SUBHEADER_FILL, fill_type="solid")
        fill_buy = PatternFill(start_color=self.BUY_FILL, end_color=self.BUY_FILL, fill_type="solid")
        fill_sell = PatternFill(start_color=self.SELL_FILL, end_color=self.SELL_FILL, fill_type="solid")
        
        font_buy = Font(name="Calibri", size=10, bold=True, color=self.BUY_FONT)
        font_sell = Font(name="Calibri", size=10, bold=True, color=self.SELL_FONT)

        thin_border_side = Side(border_style="thin", color=self.BORDER_COLOR)
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        double_bottom = Border(bottom=Side(border_style="double", color=self.NAVY_HEADER), top=thin_border_side)

        # -------------------------------------------------------------
        # 1. HEADER BANNER
        # -------------------------------------------------------------
        ws.merge_cells("A1:N2")
        title_cell = ws["A1"]
        title_cell.value = f"  {self.fund_name.upper()} — PORTFOLIO REBALANCING ORDER SHEET"
        title_cell.font = font_title
        title_cell.fill = fill_navy
        title_cell.alignment = Alignment(vertical="center")

        ws["A3"] = f"Portfolio AUM: ${self.aum:,.2f} {self.currency}  |  Execution Mandate: Black-Litterman UCITS Constrained"
        ws["A3"].font = font_bold

        # -------------------------------------------------------------
        # 2. KEY RISK & EXECUTION KPI CARDS (Row 5-7)
        # -------------------------------------------------------------
        kpis = [
            ("Expected Return (Ann.)", f"{risk_metrics.get('expected_return', 0.0):.2%}", "B5", "C5"),
            ("Portfolio Volatility", f"{risk_metrics.get('volatility', 0.0):.2%}", "D5", "E5"),
            ("Active Tracking Error", f"{risk_metrics.get('tracking_error', 0.0):.2%}", "F5", "G5"),
            ("Active Share", f"{risk_metrics.get('active_share', 0.0):.1%}", "H5", "I5"),
            ("Cornish-Fisher VaR (99%)", f"{risk_metrics.get('cf_var_99', 0.0):.2%}", "J5", "K5"),
            ("Est. Rebalance Cost", f"${risk_metrics.get('total_tx_cost', 0.0):,.2f}", "L5", "M5"),
        ]

        for label, val, top_l, btm_r in kpis:
            c_label = ws[top_l]
            c_label.value = label
            c_label.font = Font(name="Calibri", size=9, bold=True, color="555555")
            c_label.alignment = Alignment(horizontal="center")
            
            # Value cell directly below
            row_val = int(top_l[1:]) + 1
            col_letter = top_l[0]
            val_cell = ws[f"{col_letter}{row_val}"]
            val_cell.value = val
            val_cell.font = Font(name="Calibri", size=12, bold=True, color=self.NAVY_HEADER)
            val_cell.alignment = Alignment(horizontal="center")
            val_cell.fill = fill_sub

        # -------------------------------------------------------------
        # 3. TRADE EXECUTION TABLE (Row 9+)
        # -------------------------------------------------------------
        ws["A9"] = "1. TRADE ALLOCATION & EXECUTION ORDERS"
        ws["A9"].font = font_section

        headers = [
            "Ticker", "Asset Description", "Asset Class", "Last Price",
            "Current Wgt", "Current Val ($)", "Current Shares",
            "Target Wgt", "Target Val ($)", "Target Shares",
            "Wgt Delta", "Order Action", "Order Value ($)", "Shares to Trade", "Est. Cost ($)"
        ]

        start_row = 10
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        ws.row_dimensions[start_row].height = 24

        curr_row = start_row + 1
        for _, row in trades_df.iterrows():
            ws.cell(row=curr_row, column=1, value=row["ticker"]).alignment = Alignment(horizontal="center")
            ws.cell(row=curr_row, column=2, value=row["asset_name"])
            ws.cell(row=curr_row, column=3, value=row["asset_class"])
            
            c_px = ws.cell(row=curr_row, column=4, value=row["price"])
            c_px.number_format = "$#,##0.00"

            c_cw = ws.cell(row=curr_row, column=5, value=row["current_weight"])
            c_cw.number_format = "0.00%"

            c_cv = ws.cell(row=curr_row, column=6, value=row["current_value"])
            c_cv.number_format = "$#,##0"

            c_cs = ws.cell(row=curr_row, column=7, value=row["current_shares"])
            c_cs.number_format = "#,##0"

            c_tw = ws.cell(row=curr_row, column=8, value=row["target_weight"])
            c_tw.number_format = "0.00%"

            c_tv = ws.cell(row=curr_row, column=9, value=row["target_value"])
            c_tv.number_format = "$#,##0"

            c_ts = ws.cell(row=curr_row, column=10, value=row["target_shares"])
            c_ts.number_format = "#,##0"

            c_dw = ws.cell(row=curr_row, column=11, value=row["weight_delta"])
            c_dw.number_format = "+0.00%;-0.00%;0.00%"

            c_act = ws.cell(row=curr_row, column=12, value=row["action"])
            c_act.alignment = Alignment(horizontal="center")
            if row["action"] == "BUY":
                c_act.fill = fill_buy
                c_act.font = font_buy
            elif row["action"] == "SELL":
                c_act.fill = fill_sell
                c_act.font = font_sell

            c_ov = ws.cell(row=curr_row, column=13, value=row["trade_value"])
            c_ov.number_format = "$#,##0"

            c_st = ws.cell(row=curr_row, column=14, value=row["trade_shares"])
            c_st.number_format = "#,##0"

            c_tc = ws.cell(row=curr_row, column=15, value=row["tx_cost"])
            c_tc.number_format = "$#,##0.00"

            for col_i in range(1, 16):
                cell = ws.cell(row=curr_row, column=col_i)
                cell.border = cell_border
                if col_i not in [12]:  # Keep action specific font
                    cell.font = font_regular

            curr_row += 1

        # TOTALS ROW
        tot_row = curr_row
        ws.cell(row=tot_row, column=2, value="TOTAL PORTFOLIO").font = font_bold
        
        c_tot_cw = ws.cell(row=tot_row, column=5, value=f"=SUM(E{start_row+1}:E{tot_row-1})")
        c_tot_cw.number_format = "0.00%"
        c_tot_cw.font = font_bold

        c_tot_cv = ws.cell(row=tot_row, column=6, value=f"=SUM(F{start_row+1}:F{tot_row-1})")
        c_tot_cv.number_format = "$#,##0"
        c_tot_cv.font = font_bold

        c_tot_tw = ws.cell(row=tot_row, column=8, value=f"=SUM(H{start_row+1}:H{tot_row-1})")
        c_tot_tw.number_format = "0.00%"
        c_tot_tw.font = font_bold

        c_tot_tv = ws.cell(row=tot_row, column=9, value=f"=SUM(I{start_row+1}:I{tot_row-1})")
        c_tot_tv.number_format = "$#,##0"
        c_tot_tv.font = font_bold

        c_tot_ov = ws.cell(row=tot_row, column=13, value=f"=SUM(M{start_row+1}:M{tot_row-1})")
        c_tot_ov.number_format = "$#,##0"
        c_tot_ov.font = font_bold

        c_tot_tc = ws.cell(row=tot_row, column=15, value=f"=SUM(O{start_row+1}:O{tot_row-1})")
        c_tot_tc.number_format = "$#,##0.00"
        c_tot_tc.font = font_bold

        for col_i in range(1, 16):
            ws.cell(row=tot_row, column=col_i).border = double_bottom

        # -------------------------------------------------------------
        # 4. ASSET CLASS COMPLIANCE AUDIT TABLE
        # -------------------------------------------------------------
        ac_start = tot_row + 3
        ws.cell(row=ac_start, column=1, value="2. UCITS / MANDATE ASSET CLASS COMPLIANCE AUDIT").font = font_section

        ac_headers = ["Asset Class", "Current Wgt", "Target Wgt", "Mandate Min", "Mandate Max", "Compliance Status"]
        for col_idx, h in enumerate(ac_headers, start=1):
            cell = ws.cell(row=ac_start+1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center")
            cell.border = cell_border

        r_idx = ac_start + 2
        for _, ac_row in asset_class_summary.iterrows():
            ws.cell(row=r_idx, column=1, value=ac_row["asset_class"]).font = font_bold
            
            c1 = ws.cell(row=r_idx, column=2, value=ac_row["current_wgt"])
            c1.number_format = "0.00%"
            c2 = ws.cell(row=r_idx, column=3, value=ac_row["target_wgt"])
            c2.number_format = "0.00%"
            c3 = ws.cell(row=r_idx, column=4, value=ac_row["min_limit"])
            c3.number_format = "0.00%"
            c4 = ws.cell(row=r_idx, column=5, value=ac_row["max_limit"])
            c4.number_format = "0.00%"
            
            c_stat = ws.cell(row=r_idx, column=6, value=ac_row["status"])
            c_stat.alignment = Alignment(horizontal="center")
            c_stat.font = font_buy
            c_stat.fill = fill_buy

            for c_i in range(1, 7):
                ws.cell(row=r_idx, column=c_i).border = cell_border

            r_idx += 1

        # -------------------------------------------------------------
        # 5. SIGN-OFF BLOCK
        # -------------------------------------------------------------
        sign_start = r_idx + 2
        ws.cell(row=sign_start, column=1, value="Portfolio Analyst: Mohammad Ahmadian").font = font_bold
        ws.cell(row=sign_start, column=6, value="Lead Portfolio Manager Approval: _____________________").font = font_bold

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_filepath)
        logger.info(f"Excel Rebalancing Order Sheet saved to: {output_filepath}")
        return output_filepath